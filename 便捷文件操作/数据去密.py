import random
import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook('../获取捐款类信息/alumni.xlsx')
sheet = workbook.active

# 创建一个常用汉字列表，用于替换
common_chars = [
    '明', '伟', '芳', '丽', '军', '静', '鹏', '霞', '婷', '波',
    '强', '娟', '斌', '欣', '勇', '峰', '璐', '涛', '雪', '超',
    '华', '莹', '媛', '杰', '玲', '蕾', '东', '春', '浩', '颖',
    '志', '琪', '蕊', '晓', '俊', '琴', '义', '义', '凡', '燕',
    '玥', '冰', '悦', '皓', '妍', '露', '阳', '欣', '明', '达',
    '淼', '伟', '豪', '洁', '博', '鑫', '凯', '峰', '琪', '捷',
    '晗', '瑞', '梦', '航', '娜', '睿', '佳', '晓', '泽', '宁',
    '世', '皓', '瑜', '哲', '瑜', '昊', '博', '睿', '璇', '思',
    '莺', '秋', '宇', '琦', '彦', '晨', '华', '星', '莉', '珊',
    '雪', '慧', '妮', '凡', '蕙', '文', '馨', '晴', '斌', '凡'
]

# 对第2列（姓名列）进行字符替换
for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value  # 原来的姓名在第一列
    if name:
        name_length = len(name)
        # 找出不是数字的位置
        non_digit_indices = [i for i in range(name_length) if not name[i].isdigit()]

        if non_digit_indices:
            # 随机选择要替换的字的位置
            num_replacements = random.randint(1, min(2, len(non_digit_indices)))  # 最多替换两个字
            indices_to_replace = random.sample(non_digit_indices, num_replacements)

            encrypted_name = list(name)  # 将姓名转为列表以便修改
            for index in indices_to_replace:
                # 替换指定位置的字
                encrypted_name[index] = random.choice(common_chars)

            # 将加密后的姓名保存到 Excel
            sheet.cell(row=row, column=2, value=''.join(encrypted_name))

# 保存加密后的 Excel
workbook.save('encrypted_alumni.xlsx')

print("姓名加密完成，结果已保存为 'encrypted_alumni.xlsx'。")
