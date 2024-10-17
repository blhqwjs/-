import openpyxl
import random

# 打开 Excel 文件
workbook = openpyxl.load_workbook('alumni.xlsx')
sheet = workbook.active

# 汉字的 Unicode 范围
zh_start = 0x4E00  # 汉字起始
zh_end = 0x9FFF    # 汉字结束

# 对第2列（假设是姓名列）进行字符偏移加密
for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value
    if name:
        encrypted_name = []
        for char in name:
            offset = random.randint(-2, -1)  # 随机调整偏移量
            new_char = chr(ord(char) + offset)

            # 确保新字符在汉字范围内
            if zh_start <= ord(new_char) <= zh_end:
                encrypted_name.append(new_char)
            else:
                # 如果超出范围，保持原字符
                encrypted_name.append(char)

        # 将加密后的姓名保存到 Excel
        sheet.cell(row=row, column=2, value=''.join(encrypted_name))

# 保存加密后的 Excel
workbook.save('encrypted_alumni2.xlsx')

# 此方法虽然可行，但整体影响可读性